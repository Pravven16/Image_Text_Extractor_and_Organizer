import cv2
from flask import Flask, render_template, jsonify, request, send_from_directory
from PIL import Image
import pytesseract
from geopy.geocoders import Nominatim
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)
app.static_folder = 'static'  # Set the static folder explicitly

geolocator = Nominatim(user_agent="Flask(__name__)")

@app.route('/')
def index():
    return render_template('index.html')

def preprocess_image(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    thresholded = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    return thresholded

def tesseract(image):
    path_to_tesseract = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    pytesseract.pytesseract.tesseract_cmd = path_to_tesseract

    thresholded = preprocess_image(image)

    image_rgb = cv2.cvtColor(thresholded, cv2.COLOR_GRAY2RGB)

    results = pytesseract.image_to_boxes(image_rgb)

    extracted_text = ""
    for box in results.split('\n'):
        box = box.split(' ')
        if len(box) >= 6:  # Ensure valid box format
            char = box[0]
            extracted_text += char

    return extracted_text

@app.route('/capture', methods=['POST'])
def capture():
    camera = cv2.VideoCapture(0)
    current_datetime = datetime.now()
    image_filename = current_datetime.strftime("%Y%m%d_%H%M%S") + ".png"

    while True:
        _, image = camera.read()
        cv2.imshow("Capturing", image)
        if cv2.waitKey(1) & 0xFF == ord('s'):
            image_path = os.path.join("static", image_filename)
            cv2.imwrite(image_path, image)
            break

    camera.release()
    cv2.destroyAllWindows()

    name = tesseract(image)
    ip = request.remote_addr
    location = geolocator.geocode(ip)

    excel_file = "output.xlsx"
    sheet_name = "Sheet"

    if not os.path.exists(excel_file):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["Date_Time", "Name", "Place", "Image_filename"])
    else:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[sheet_name]

    date_time = datetime.now()
    place = location.address

    worksheet.append([date_time, name, place, image_filename])

    for row in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    worksheet.column_dimensions['A'].width = 20

    workbook.save(excel_file)

    return jsonify({
        "message": "Image captured successfully",
        "extracted_text": name,
        "photo": image_filename
    })

@app.route('/search', methods=['POST'])
def search():
    search_text = request.json.get('search_text', '')
    
    search_results = search_excel_by_name(search_text)

    return jsonify(search_results)

def search_excel_by_name(search_text):
    excel_file = "output.xlsx"
    
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        search_results = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date_time, name, place, photo = row
            if name is not None and search_text.lower() in name.lower():
                search_results.append({
                    "date_time": date_time.strftime('%Y-%m-%d %H:%M'),
                    "name": name,
                    "place": place,
                    "photo": photo
                })

        return search_results
    else:
        return []

if __name__ == '__main__':
    app.run(debug=True)
